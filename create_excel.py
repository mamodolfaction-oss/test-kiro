from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

# Read data from file
with open("/projects/sandbox/test-kiro/data_clean.tsv", "r") as f:
    content = f.read()

lines = [l for l in content.strip().split('\n') if l.strip()]
print(f"Total data rows: {len(lines)}")

# Parse: X1=col0(dept), X2=col1(angkatan)
# X1: 1=Teknik, 2=Bisnis, 3=Kesehatan
# X2: 1=2023, 2=2024, 3=2025
dept_map = {1: "Teknik", 2: "Bisnis", 3: "Kesehatan"}
ang_map = {1: 2023, 2: 2024, 3: 2025}

# Target responden per cell (proporsional n=128)
target = {
    (2023, "Teknik"): 7, (2023, "Bisnis"): 17, (2023, "Kesehatan"): 17,
    (2024, "Teknik"): 8, (2024, "Bisnis"): 19, (2024, "Kesehatan"): 20,
    (2025, "Teknik"): 7, (2025, "Bisnis"): 17, (2025, "Kesehatan"): 16,
}


# Parse all rows
all_data = []
for i, line in enumerate(lines):
    cols = line.split('\t')
    x1 = int(cols[0]) if cols[0].strip() else 0
    x2 = int(cols[1]) if cols[1].strip() else 0
    dept = dept_map.get(x1, "Unknown")
    ang = ang_map.get(x2, 0)
    all_data.append({
        'row_idx': i+1,
        'dept': dept,
        'angkatan': ang,
        'raw': cols,
        'assigned': None  # will be 'responden' or 'pilot'
    })

# Count per cell
from collections import defaultdict
cell_count = defaultdict(list)
for d in all_data:
    key = (d['angkatan'], d['dept'])
    cell_count[key].append(d)

# Assign: first N rows per cell = responden, rest = pilot
for key, items in cell_count.items():
    tgt = target.get(key, 0)
    for i, item in enumerate(items):
        if i < tgt:
            item['assigned'] = 'responden'
        else:
            item['assigned'] = 'pilot'

# Verify
resp_count = sum(1 for d in all_data if d['assigned'] == 'responden')
pilot_count = sum(1 for d in all_data if d['assigned'] == 'pilot')
print(f"Responden: {resp_count}, Pilot: {pilot_count}")


# Column headers
col_names = ["X1","X2","X3","X4","X5","L1","L2","L3","L4","L5","L6",
             "M1","M2","M3","M4","I2","I3","I4","X10",
             "W1","W2","W3","H1","H2","X11a","X11b","X12a","X12b","X12c"]

# Create Excel
wb = Workbook()

# Sheet 1: Responden
ws_resp = wb.active
ws_resp.title = "Responden (n=128)"
headers_resp = ["No"] + col_names + ["Departemen", "Angkatan"]
for col, h in enumerate(headers_resp, 1):
    ws_resp.cell(row=1, column=col, value=h).font = Font(bold=True)

resp_rows = [d for d in all_data if d['assigned'] == 'responden']
for i, d in enumerate(resp_rows):
    row = i + 2
    ws_resp.cell(row=row, column=1, value=i+1)
    for j, val in enumerate(d['raw'][:len(col_names)]):
        v = val.strip() if val.strip() else ""
        try:
            v = int(v)
        except:
            pass
        ws_resp.cell(row=row, column=j+2, value=v)
    ws_resp.cell(row=row, column=len(col_names)+2, value=d['dept'])
    ws_resp.cell(row=row, column=len(col_names)+3, value=d['angkatan'])

# Sheet 2: Pilot Test
ws_pilot = wb.create_sheet("Pilot Test (n=30)")
headers_pilot = ["No"] + col_names + ["Departemen", "Angkatan"]
for col, h in enumerate(headers_pilot, 1):
    ws_pilot.cell(row=1, column=col, value=h).font = Font(bold=True)

pilot_rows = [d for d in all_data if d['assigned'] == 'pilot']
for i, d in enumerate(pilot_rows):
    row = i + 2
    ws_pilot.cell(row=row, column=1, value=i+1)
    for j, val in enumerate(d['raw'][:len(col_names)]):
        v = val.strip() if val.strip() else ""
        try:
            v = int(v)
        except:
            pass
        ws_pilot.cell(row=row, column=j+2, value=v)
    ws_pilot.cell(row=row, column=len(col_names)+2, value=d['dept'])
    ws_pilot.cell(row=row, column=len(col_names)+3, value=d['angkatan'])


# Sheet 3: Distribusi Sampel
ws_dist = wb.create_sheet("Distribusi Sampel")
ws_dist.cell(row=1, column=1, value="DISTRIBUSI DATA").font = Font(bold=True, size=14)

# Tabel responden
ws_dist.cell(row=3, column=1, value="RESPONDEN (n=128)").font = Font(bold=True)
for col, h in enumerate(["Angkatan", "Teknik", "Bisnis", "Kesehatan", "Total"], 1):
    ws_dist.cell(row=4, column=col, value=h).font = Font(bold=True)
resp_data = [(2023,7,17,17,41),(2024,8,19,20,47),(2025,7,17,16,40),("Total",22,53,53,128)]
for i, row_data in enumerate(resp_data):
    for j, val in enumerate(row_data):
        ws_dist.cell(row=5+i, column=j+1, value=val)

# Tabel pilot
ws_dist.cell(row=11, column=1, value="PILOT TEST (n=30)").font = Font(bold=True)
for col, h in enumerate(["Angkatan", "Teknik", "Bisnis", "Kesehatan", "Total"], 1):
    ws_dist.cell(row=12, column=col, value=h).font = Font(bold=True)
pilot_data = [(2023,0,3,2,5),(2024,6,5,0,11),(2025,12,1,1,14),("Total",18,9,3,30)]
for i, row_data in enumerate(pilot_data):
    for j, val in enumerate(row_data):
        ws_dist.cell(row=13+i, column=j+1, value=val)

# Tabel total
ws_dist.cell(row=19, column=1, value="TOTAL DATA CLEAN (n=158)").font = Font(bold=True)
for col, h in enumerate(["Angkatan", "Teknik", "Bisnis", "Kesehatan", "Total"], 1):
    ws_dist.cell(row=20, column=col, value=h).font = Font(bold=True)
total_data = [(2023,7,20,19,46),(2024,14,24,20,58),(2025,19,18,17,54),("Total",40,62,56,158)]
for i, row_data in enumerate(total_data):
    for j, val in enumerate(row_data):
        ws_dist.cell(row=21+i, column=j+1, value=val)

# Sheet 4: Perhitungan Proporsi dari Pilot
ws_prop = wb.create_sheet("Perhitungan Proporsi")
ws_prop.cell(row=1, column=1, value="PERHITUNGAN PROPORSI DARI PILOT TEST").font = Font(bold=True, size=14)
ws_prop.cell(row=3, column=1, value="Variabel Y (Loyalitas) = L1+L2+L3+L4+L5+L6")
ws_prop.cell(row=4, column=1, value="Skoring: 1=Tidak Setuju, 2=Cukup Setuju, 3=Setuju, 4=Sangat Setuju")
ws_prop.cell(row=5, column=1, value="Kategorisasi: Rendah(6-12), Sedang(13-18), Tinggi(19-24)")

# Calculate Y for pilot rows
for col, h in enumerate(["No","L1","L2","L3","L4","L5","L6","Skor_Y","Kategori"], 1):
    ws_prop.cell(row=7, column=col, value=h).font = Font(bold=True)

rendah = sedang = tinggi = 0
for i, d in enumerate(pilot_rows):
    row = i + 8
    ws_prop.cell(row=row, column=1, value=i+1)
    skor_y = 0
    for j in range(6):  # L1-L6 = cols index 5-10
        val = d['raw'][5+j].strip() if len(d['raw']) > 5+j else ""
        try:
            v = int(val)
        except:
            v = 0
        ws_prop.cell(row=row, column=j+2, value=v)
        skor_y += v
    ws_prop.cell(row=row, column=8, value=skor_y)
    if 6 <= skor_y <= 12:
        kat = "Rendah"
        rendah += 1
    elif 13 <= skor_y <= 18:
        kat = "Sedang"
        sedang += 1
    else:
        kat = "Tinggi"
        tinggi += 1
    ws_prop.cell(row=row, column=9, value=kat)

# Summary
sum_row = 8 + len(pilot_rows) + 1
ws_prop.cell(row=sum_row, column=1, value="RINGKASAN").font = Font(bold=True)
ws_prop.cell(row=sum_row+1, column=1, value="Rendah")
ws_prop.cell(row=sum_row+1, column=2, value=rendah)
ws_prop.cell(row=sum_row+1, column=3, value=f"{rendah/30:.4f}")
ws_prop.cell(row=sum_row+2, column=1, value="Sedang")
ws_prop.cell(row=sum_row+2, column=2, value=sedang)
ws_prop.cell(row=sum_row+2, column=3, value=f"{sedang/30:.4f}")
ws_prop.cell(row=sum_row+3, column=1, value="Tinggi")
ws_prop.cell(row=sum_row+3, column=2, value=tinggi)
ws_prop.cell(row=sum_row+3, column=3, value=f"{tinggi/30:.4f}")

print(f"Pilot Y categories: Rendah={rendah}, Sedang={sedang}, Tinggi={tinggi}")

# Save
wb.save("/projects/sandbox/test-kiro/perhitungan_sampel.xlsx")
print("Excel saved!")
